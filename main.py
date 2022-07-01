import dataclasses
import openpyxl as op
import os

# DTOクラス定義
@dataclasses.dataclass
class CreateCombinationType(dict):
    group: str
    categories: list[str]
    result: list[tuple]
 
class CreateCombination:
  """
  @input
  input = [
    {
      'group': 'group1',
      'categories': ['category1', 'category2']
    },
    {
      'group': 'group2',
      'categories': ['category1', 'category2']
    }
  ]
  @output
  output = [
    {
      'group': 'group1',
      'categories': ['category1', 'category2'],
      'result': [(0, '○'), (0, '○'), (1, '○'), (1, '○')]
    },
    {
      'group': 'group2',
      'categories': ['category1', 'category2'],
      'result': [(0, '○'), (1, '○'), (0, '○'), (1, '○')]
    }
  ]
  """
  def __init__(self) -> None:
      pass

  def pattern(self, dataList: list[CreateCombinationType], deepCount:int = 0, props:int = 1):
    deepDiff = len(dataList) - deepCount
    if (deepDiff == 1):
      return props
    for index in range(deepCount+1, len(dataList)):
      props *= len(dataList[index].get('categories'))
    return props

  def createComb(self, dataList: list[CreateCombinationType], deepCount:int = 0):
    result = dataList[deepCount].get('result') or list()
    for index, _ in enumerate (dataList[deepCount].get('categories')):
      loopCount = self.pattern(dataList, deepCount)
      for _ in range(loopCount):
        result.append((index , '○'))
      if not (deepCount == len(dataList) - 1):
        self.createComb(dataList, deepCount = deepCount+1)
    dataList[deepCount]['result'] = result
    return dataList

  def createXlsx(self, dataList:list[CreateCombinationType]):
    # print(dataList)
    wb = op.Workbook()
    acst = wb.active
    path = os.path.dirname(os.path.abspath(__file__))
    fileName = os.path.normpath(os.path.join(path, 'sample.xlsx'))

    gCol = 1
    cCol = 2
    pCol = 3
    row = 1
    for data in dataList:
      group = data.get('group')
      acst.cell(row=row, column=gCol, value=group)
      for cIndex, category in enumerate(data.get('categories')):
        acst.cell(row=row + cIndex, column=cCol, value=category)
      for pIndex, res in enumerate(data.get('result')):
        posY, value = res
        acst.cell(row=row + posY, column=pCol+pIndex, value=value)
      row += len(data.get('categories'))

    wb.save(fileName)
    pass


  def run(self):
    input:CreateCombinationType = [
      {
       'group': '新・旧サーバ',
       'categories': ['新サバ', '旧サバ']
      },
      {
        'group': '所属グループ',
        'categories': ['100', '200', '300']
      },
      {
        'group': 'ロール',
        'categories': ['一般', 'リーダー', 'マネージャー', 'エリアマネージャー']
      },
      {
        'group': '作業内容',
        'categories': ['作業A', '作業B', '作業C', '作業D', '作業E', '作業F', '作業G']
      }
    ]
    output = self.createComb(input)
    self.createXlsx(output)

cc = CreateCombination()
cc.run()